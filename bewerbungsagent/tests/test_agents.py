import json
import os
import sys
import importlib.util
import tempfile
import pytest
from unittest.mock import patch, MagicMock

# Agenten per importlib laden (wegen numerischen Dateinamen)
AGENTS_ORDNER = os.path.join(os.path.dirname(__file__), "..", "agents")


def _lade_modul(name, dateiname):
    pfad = os.path.join(AGENTS_ORDNER, dateiname)
    spec = importlib.util.spec_from_file_location(name, pfad)
    modul = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modul)
    return modul


skill_extractor = _lade_modul("skill_extractor", "02_skill_extractor.py")

# Excel-Handler direkt importieren
sys.path.append(os.path.join(os.path.dirname(__file__), ".."))
from utils.excel_handler import add_bewerbung, EXCEL_PFAD, HEADERS
import openpyxl


# =====================================================================
# Tests für update_profil
# =====================================================================

def test_neuer_skill_wird_hinzugefuegt(tmp_path):
    """Prüft, dass ein neuer Skill korrekt ins Profil geschrieben wird."""

    # Temporäres Profil verwenden
    test_profil_pfad = str(tmp_path / "profil.json")

    with patch.object(skill_extractor, "PROFIL_PFAD", test_profil_pfad):
        neue_skills = {
            "hard_skills": ["Python"],
            "soft_skills": [],
            "tools": [],
            "erfahrungslevel": {"Python": 3}
        }
        skill_extractor.update_profil(neue_skills)

        # Profil laden und prüfen
        with open(test_profil_pfad, "r", encoding="utf-8") as f:
            profil = json.load(f)

    skill_namen = [s["name"] for s in profil["skills"]]
    assert "Python" in skill_namen, "Python sollte im Profil sein"


def test_duplikat_behaelt_hoeheren_level(tmp_path):
    """Prüft, dass bei doppeltem Skill der höhere Level behalten wird."""

    test_profil_pfad = str(tmp_path / "profil.json")

    with patch.object(skill_extractor, "PROFIL_PFAD", test_profil_pfad):
        # Ersten Skill mit Level 2 einfügen
        skill_extractor.update_profil({
            "hard_skills": ["Python"],
            "soft_skills": [],
            "tools": [],
            "erfahrungslevel": {"Python": 2}
        })

        # Denselben Skill mit Level 4 einfügen
        skill_extractor.update_profil({
            "hard_skills": ["Python"],
            "soft_skills": [],
            "tools": [],
            "erfahrungslevel": {"Python": 4}
        })

        with open(test_profil_pfad, "r", encoding="utf-8") as f:
            profil = json.load(f)

    python_skills = [s for s in profil["skills"] if s["name"] == "Python"]
    assert len(python_skills) == 1, "Python sollte nur einmal vorkommen"
    assert python_skills[0]["level"] == 4, "Level sollte 4 (der höhere) sein"


def test_duplikat_behaelt_niedrigeren_level_unveraendert(tmp_path):
    """Prüft, dass ein niedrigerer Level einen bestehenden höheren Level nicht überschreibt."""

    test_profil_pfad = str(tmp_path / "profil.json")

    with patch.object(skill_extractor, "PROFIL_PFAD", test_profil_pfad):
        # Erst Level 5 einfügen
        skill_extractor.update_profil({
            "hard_skills": ["Java"],
            "soft_skills": [],
            "tools": [],
            "erfahrungslevel": {"Java": 5}
        })

        # Dann Level 2 einfügen — sollte Level 5 NICHT überschreiben
        skill_extractor.update_profil({
            "hard_skills": ["Java"],
            "soft_skills": [],
            "tools": [],
            "erfahrungslevel": {"Java": 2}
        })

        with open(test_profil_pfad, "r", encoding="utf-8") as f:
            profil = json.load(f)

    java_skill = next(s for s in profil["skills"] if s["name"] == "Java")
    assert java_skill["level"] == 5, "Level 5 sollte erhalten bleiben"


# =====================================================================
# Tests für extract_from_text
# =====================================================================

def test_extract_from_text_gibt_dict_mit_erwarteten_keys_zurueck(tmp_path):
    """Prüft, dass extract_from_text ein Dict mit den erwarteten Schlüsseln zurückgibt."""

    test_profil_pfad = str(tmp_path / "profil.json")

    # Ollama-Aufruf mocken
    mock_antwort = json.dumps({
        "hard_skills": ["Python", "Django"],
        "soft_skills": ["Teamarbeit"],
        "tools": ["Git"],
        "erfahrungslevel": {"Python": 3, "Django": 2}
    })

    with patch.object(skill_extractor, "PROFIL_PFAD", test_profil_pfad):
        # ask_ollama direkt im geladenen Modul patchen (from-Import hat eigene Referenz)
        with patch.object(skill_extractor, "ask_ollama", return_value=mock_antwort):
            ergebnis = skill_extractor.extract_from_text("Ich entwickle mit Python und Django.")

    assert ergebnis is not None, "Ergebnis sollte nicht None sein"
    assert "hard_skills" in ergebnis, "Schlüssel 'hard_skills' fehlt"
    assert "soft_skills" in ergebnis, "Schlüssel 'soft_skills' fehlt"
    assert "tools" in ergebnis, "Schlüssel 'tools' fehlt"
    assert "erfahrungslevel" in ergebnis, "Schlüssel 'erfahrungslevel' fehlt"


# =====================================================================
# Tests für match_profile
# =====================================================================

def test_match_score_liegt_zwischen_0_und_100(tmp_path):
    """Prüft, dass der Match-Score zwischen 0 und 100 liegt."""

    test_profil_pfad = str(tmp_path / "profil.json")

    # Profil mit einem Skill anlegen
    test_profil = {
        "version": 1,
        "nutzer_id": "user_001",
        "skills": [{"name": "Python", "kategorie": "Hard Skill", "level": 3, "quelle": "test", "zuletzt_aktualisiert": "2026-05-30"}],
        "erfahrungen": [],
        "letztes_update": "2026-05-30"
    }
    with open(test_profil_pfad, "w", encoding="utf-8") as f:
        json.dump(test_profil, f)

    mock_antwort = json.dumps({
        "match_score": 75,
        "vorhandene_skills": ["Python"],
        "fehlende_skills": ["Docker"],
        "lernbare_skills": ["Kubernetes"]
    })

    job = {"titel": "Python Entwickler", "unternehmen": "Test GmbH", "beschreibung": "Python Docker Kubernetes"}

    # application_agent importieren und testen
    application_agent = _lade_modul("application_agent", "03_application_agent.py")

    with patch.object(application_agent, "PROFIL_PFAD", test_profil_pfad):
        with patch.object(application_agent, "ask_ollama", return_value=mock_antwort):
            ergebnis = application_agent.match_profile(job)

    assert ergebnis is not None, "match_profile sollte ein Ergebnis zurückgeben"
    assert 0 <= ergebnis["match_score"] <= 100, "Match-Score muss zwischen 0 und 100 liegen"


def test_match_score_wird_auf_gueltigen_bereich_begrenzt(tmp_path):
    """Prüft, dass ein Score > 100 auf 100 begrenzt wird."""

    test_profil_pfad = str(tmp_path / "profil.json")

    test_profil = {
        "version": 1,
        "nutzer_id": "user_001",
        "skills": [{"name": "Python", "kategorie": "Hard Skill", "level": 5, "quelle": "test", "zuletzt_aktualisiert": "2026-05-30"}],
        "erfahrungen": [],
        "letztes_update": "2026-05-30"
    }
    with open(test_profil_pfad, "w", encoding="utf-8") as f:
        json.dump(test_profil, f)

    # LLM gibt ungültigen Score zurück
    mock_antwort = json.dumps({"match_score": 150, "vorhandene_skills": [], "fehlende_skills": [], "lernbare_skills": []})
    job = {"titel": "Test Job", "beschreibung": "Python"}

    application_agent = _lade_modul("application_agent", "03_application_agent.py")

    with patch.object(application_agent, "PROFIL_PFAD", test_profil_pfad):
        with patch.object(application_agent, "ask_ollama", return_value=mock_antwort):
            ergebnis = application_agent.match_profile(job)

    assert ergebnis["match_score"] == 100, "Score > 100 soll auf 100 begrenzt werden"


# =====================================================================
# Tests für add_bewerbung (Excel)
# =====================================================================

def test_excel_datei_wird_erstellt_und_hat_korrekten_header(tmp_path):
    """Prüft, dass add_bewerbung eine Excel-Datei mit korrekten Spaltenüberschriften erstellt."""

    test_excel_pfad = str(tmp_path / "bewerbungen.xlsx")

    with patch("utils.excel_handler.EXCEL_PFAD", test_excel_pfad):
        add_bewerbung("Python Entwickler", "Test GmbH", 85, "Gesendet", "https://example.com")

    assert os.path.exists(test_excel_pfad), "Excel-Datei sollte erstellt worden sein"

    wb = openpyxl.load_workbook(test_excel_pfad)
    ws = wb.active

    # Header-Zeile prüfen
    header = [ws.cell(1, col).value for col in range(1, len(HEADERS) + 1)]
    assert header == HEADERS, f"Header stimmt nicht überein. Erwartet: {HEADERS}, Erhalten: {header}"


def test_excel_bewerbung_wird_als_zeile_gespeichert(tmp_path):
    """Prüft, dass die Bewerbungsdaten korrekt in die Excel-Datei geschrieben werden."""

    test_excel_pfad = str(tmp_path / "bewerbungen.xlsx")

    with patch("utils.excel_handler.EXCEL_PFAD", test_excel_pfad):
        add_bewerbung("Data Scientist", "KI Corp", 92, "Gesendet", "https://example.com/job/123")

    wb = openpyxl.load_workbook(test_excel_pfad)
    ws = wb.active

    # Erste Datenzeile (Zeile 2, da Zeile 1 der Header ist)
    assert ws["A2"].value == "Data Scientist", "Job-Titel in Spalte A falsch"
    assert ws["B2"].value == "KI Corp", "Unternehmen in Spalte B falsch"
    assert ws["C2"].value == 92, "Match-Score in Spalte C falsch"
    assert ws["G2"].value == "https://example.com/job/123", "URL in Spalte G falsch"
